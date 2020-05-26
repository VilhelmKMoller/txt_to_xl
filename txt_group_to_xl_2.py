#!/usr/bin/python
"""
Program extracts all data from txt files placed in folder "test name".
The files are exported to a excel file (xlsx), with the name specified by the user in the command prompt

# todo add GUI så user can decide the fils they want to add
# alt + cmd + t gives surround with statemetn
"""
import sys
import os
from openpyxl import Workbook



###################################
class OpenFile:
    def __init__(self, first_lst, third_lst, file_count):
        self.first_lst = first_lst
        self.third_lst = third_lst
        self.PREFIX = "test name/"
        self.SUFIX = '.txt'

    # open file
    def get_file_lst(self):
        # go into test folder and extract all files ending wtih '.txt'
        input_files = [f for f in os.listdir(self.PREFIX) if f.endswith(self.SUFIX)]

        # control if there are any files in the input_file vector
        if len(input_files) == 0:
            print("ERROR: There is no .txt files in the folder ")
            sys.stdout.close()
        return input_files

    def open_file(self, file, file_count):
        file_count += 1

        with open(self.PREFIX + file, 'r') as infile:
            for line in infile:
                # if line starts with First: or Third: then extrat next number that comes after
                # split line into vector and loop over vector to First: and Third:
                line_load = line.split(',')
                self.load_to_list(line_load, self.first_lst, self.third_lst)

            self.ad_info(self.first_lst, self.third_lst, file)
        return self.first_lst, self.third_lst

    def load_to_list(self, line_load, first_lst, third_lst):
        # maybe numpy.
        for num in range(0, len(line_load)):
            # remove wide spaces
            value = line_load[num].strip(' ')
            if value == 'First:':
                self.append_to_lst(first_lst, line_load, num)  # add self.append_to_list
            if value == 'Third:':
                self.append_to_lst(third_lst, line_load, num)  # add self.append_to_list
        return first_lst, third_lst

    # def add append_to_lst here.
    def append_to_lst(self, lst, line, num):
        if isinstance(line[num + 1], float) == False:
            lst.append(line[num + 1].strip(' '))
        elif isinstance(line[num + 2], float) == False:
            lst.append(line[num + 2].strip(' '))
        return lst

    def ad_info(self, first_lst, third_lst, file):
        # add 'First' and 'Third' to first line in list
        first_lst.insert(0, "First")
        third_lst.insert(0, "Third")
        # add name of old txt file to top of first_lst
        first_lst.insert(0, file)
        third_lst.insert(0, "")
        return first_lst, third_lst

class SaveToWorkbook:
    def __init__(self):
        self.workbook = Workbook()
        self.sheet = self.workbook.active
        self.SUFIX = ".xlsx"
        ## load data into excel workbook
        # first and third lst can have varying lengths

    def save_workbook(self, first_lst, third_lst, file_count):
        # open up excel file that you want to read into
        # load data into excel workbook
        self.cell_load(first_lst, third_lst, file_count, self.sheet)
        return

    def cell_load(self, list_1, list_2, file_count, worksheet):
        # test if the value after First is a float, else take next value
        for cell in range(0, len(list_1)):
            worksheet.cell(row=cell + 1, column=file_count * 2 - 1).value = list_1[cell]
        for cell in range(0, len(list_2)):
            worksheet.cell(row=cell + 1, column=file_count * 2).value = list_2[cell]
        return

    def get_name(self, input_files, file_count):
        if len(input_files) == file_count:
            # ask user for file name
            input_raw = input("Please enter file name: ")
            # check if nmae contains sufix, else add it
            if not input_raw.endswith(self.SUFIX):
                out_file = input_raw + self.SUFIX

            return self.check_name(out_file)

    def check_name(self, out_file):
        # if a file with the same name exists then add a number
        if os.path.isfile(out_file):
            number_of_files = len([f for f in os.listdir() if f.startswith(out_file)])
            new_file_name = out_file + "(" + str(number_of_files) + ")" + self.SUFIX
            self.workbook.save(filename=new_file_name)
        else:
            # when all files have been loaded then save workbook
            self.workbook.save(filename=out_file)
        return


################################### Run code #############

# initialize objects
test_extract = OpenFile([], [], 0)
save_to_file = SaveToWorkbook()
# for loop:

input_files = test_extract.get_file_lst()

file_count = 0
for file in input_files:
    # print('test_extract.open_file(file)', len(test_extract.open_file(file)))
    file_count += 1
    first_lst, third_lst = test_extract.open_file(file, file_count)

    save_to_file.save_workbook(first_lst, third_lst, file_count)
    save_to_file.get_name(input_files, file_count)

    # clean list so they are ready for use in top of loop
    test_extract.__init__([], [], file_count)

