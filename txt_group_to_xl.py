#!/usr/bin/python
"""
Program extracts all data from txt files placed in folder "test name".
The files are exported to a excel file (xlsx), with the name specified by the user in the command prompt
#todo add comments
# todo check that all the output lines match the real line lengths
# todo add GUI så user can decide the fils they want to add
# alt + cmd + t gives surround with statemetn
"""
import sys
import os
from openpyxl import Workbook

##################################################################################
def append_to_lst(lst, line):
    if isinstance(line[num + 1], float) == False:
        lst.append(line[num + 1].strip(' '))
    elif isinstance(line[num + 2], float) == False:
        lst.append(line[num + 2].strip(' '))
    return lst

def cell_load(list_1, list_2, file_count, worksheet):
    # test if the value after First is a float, else take next value
    for cell in range(0, len(list_1)):
        worksheet.cell(row=cell + 1, column=file_count * 2 - 1).value = list_1[cell]
    for cell in range(0, len(list_2)):
        worksheet.cell(row=cell + 1, column=file_count * 2).value = list_2[cell]
    return
##############################################################################

# go into test folder and extract all files ending wtih '.txt'
prefix = "test name/"
input_file = [f for f in os.listdir(prefix) if f.endswith('.txt')]

# control if there are any files in the input_file vector
if len(input_file) == 0:
    print("ERROR: There is no .txt files in the folder ")
    sys.stdout.close()

drop_list = [], first_lst = [], third_lst = [], file_count = 0
for file in input_file:
    # count files to ensure they are placed correctly in the new csv file
    file_count += 1

    # open file and read line by line.
    with open("test name/" + file, 'r') as infile:
        for line in infile:
            # if line starts with First: or Third: then extrat next number that comes after
            # split line into vector and loop over vector to First: and Third:
            line_load = line.split(',')
            for num in range(0, len(line_load)):
                # remove wide spaces
                value = line_load[num].strip(' ')
                if value == 'First:':
                    append_to_lst(first_lst, line_load)
                if value == 'Third:':
                    append_to_lst(third_lst, line_load)

    # add 'First' and 'Third' to first line in list
    first_lst.insert(0, "First")
    third_lst.insert(0, "Third")
    # add name of old txt file to top of first_lst
    first_lst.insert(0, file)
    third_lst.insert(0, "")

    # open up excel file that you want to read into
    if file_count == 1:
        workbook = Workbook()
        sheet = workbook.active
        ## load data into excel workbook
        # first and third lst can have varying lengths
        cell_load(first_lst, third_lst, file_count, sheet)
    else:
        # load data into excel workbook
        cell_load(first_lst, third_lst, file_count, sheet)

    if len(input_file) == file_count:
        # ask user for file name
        input_raw = input("Please enter file name: ")
        # check if nmae contains sufix, else add it
        if not input_raw.endswith(".xlsx"):
            out_file = input_raw + ".xlsx"

        # if a file with the same name exists then add a number
        if os.path.isfile(out_file):
            number_of_files = len([f for f in os.listdir() if f.startswith(out_file)])
            new_file_name = out_file + "(" + str(number_of_files) + ").xlsx"
            workbook.save(filename=new_file_name)
        else:
            # when all files have been loaded then save workbook
            workbook.save(filename=out_file)
    # clean list so they are ready for use in top of loop
    first_lst = [], third_lst = []
